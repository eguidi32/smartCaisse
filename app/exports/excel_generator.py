"""
Générateur de fichiers Excel pour SmartCaisse
Utilise openpyxl pour créer des documents Excel
"""
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Générateur de documents Excel pour SmartCaisse"""
    
    # Styles
    HEADER_FILL = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TOTAL_FILL = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=11)
    BORDER = Border(
        left=Side(style='thin', color='bdc3c7'),
        right=Side(style='thin', color='bdc3c7'),
        top=Side(style='thin', color='bdc3c7'),
        bottom=Side(style='thin', color='bdc3c7')
    )
    
    def __init__(self, title="Document SmartCaisse"):
        self.title = title
    
    def _apply_header_style(self, ws, row):
        """Applique le style d'en-tête à une ligne"""
        for cell in ws[row]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
    
    def _apply_data_style(self, ws, start_row, end_row):
        """Applique le style de données aux lignes"""
        for row_idx in range(start_row, end_row + 1):
            for cell in ws[row_idx]:
                cell.border = self.BORDER
                cell.alignment = Alignment(vertical='center')
    
    def _apply_total_style(self, ws, row):
        """Applique le style de total à une ligne"""
        for cell in ws[row]:
            cell.fill = self.TOTAL_FILL
            cell.font = self.TOTAL_FONT
            cell.border = self.BORDER
    
    def _auto_width(self, ws):
        """Ajuste automatiquement la largeur des colonnes"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_title_row(self, ws, title, num_cols):
        """Ajoute une ligne de titre"""
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14, color="2c3e50")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=10, color="7f8c8d")
        ws['A2'].alignment = Alignment(horizontal='center')
    
    def generate_inventory_excel(self, products, user_name=""):
        """Génère un fichier Excel de l'inventaire"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventaire"
        
        # Titre
        title = f"Inventaire de {user_name}" if user_name else "Inventaire complet"
        self._add_title_row(ws, title, 6)
        
        # En-têtes
        headers = ['Produit', 'SKU', 'Catégorie', 'Prix (FCFA)', 'Stock', 'Valeur (FCFA)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        self._apply_header_style(ws, 4)
        
        # Données
        total_value = 0
        row = 5
        for product in products:
            stock = product.current_stock
            value = stock * product.price
            total_value += value
            
            ws.cell(row=row, column=1, value=product.name)
            ws.cell(row=row, column=2, value=product.sku or '-')
            ws.cell(row=row, column=3, value=product.category.name if product.category else '-')
            ws.cell(row=row, column=4, value=product.price)
            ws.cell(row=row, column=5, value=stock)
            ws.cell(row=row, column=6, value=value)
            row += 1
        
        self._apply_data_style(ws, 5, row - 1)
        
        # Total
        ws.cell(row=row, column=5, value='TOTAL:')
        ws.cell(row=row, column=6, value=total_value)
        self._apply_total_style(ws, row)
        
        self._auto_width(ws)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_debts_excel(self, clients, user_name=""):
        """Génère un fichier Excel des dettes clients"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Dettes"
        
        # Titre
        title = f"Dettes clients de {user_name}" if user_name else "Liste des dettes"
        self._add_title_row(ws, title, 5)
        
        # En-têtes
        headers = ['Client', 'Téléphone', 'Total dettes (FCFA)', 'Total payé (FCFA)', 'Reste à payer (FCFA)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        self._apply_header_style(ws, 4)
        
        # Données
        total_dettes = 0
        total_paye = 0
        total_restant = 0
        row = 5
        
        for client in clients:
            total_dettes += client.total_dette
            total_paye += client.total_paye
            total_restant += client.total_restant
            
            ws.cell(row=row, column=1, value=client.nom)
            ws.cell(row=row, column=2, value=client.telephone)
            ws.cell(row=row, column=3, value=client.total_dette)
            ws.cell(row=row, column=4, value=client.total_paye)
            ws.cell(row=row, column=5, value=client.total_restant)
            row += 1
        
        self._apply_data_style(ws, 5, row - 1)
        
        # Total
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=3, value=total_dettes)
        ws.cell(row=row, column=4, value=total_paye)
        ws.cell(row=row, column=5, value=total_restant)
        self._apply_total_style(ws, row)
        
        self._auto_width(ws)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_transactions_excel(self, transactions, user_name="", period=""):
        """Génère un fichier Excel des transactions"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Titre
        title = f"Transactions de {user_name}" if user_name else "Transactions"
        if period:
            title += f" - {period}"
        self._add_title_row(ws, title, 5)
        
        # En-têtes
        headers = ['Date', 'Type', 'Description', 'Catégorie', 'Montant (FCFA)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        self._apply_header_style(ws, 4)
        
        # Données
        total_recettes = 0
        total_depenses = 0
        row = 5
        
        for t in transactions:
            if t.type == 'recette':
                total_recettes += t.amount
                montant = t.amount
            else:
                total_depenses += t.amount
                montant = -t.amount
            
            ws.cell(row=row, column=1, value=t.date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value='Recette' if t.type == 'recette' else 'Dépense')
            ws.cell(row=row, column=3, value=t.description)
            ws.cell(row=row, column=4, value=t.category or '-')
            ws.cell(row=row, column=5, value=montant)
            row += 1
        
        self._apply_data_style(ws, 5, row - 1)
        
        # Totaux
        solde = total_recettes - total_depenses
        ws.cell(row=row, column=4, value='Recettes:')
        ws.cell(row=row, column=5, value=total_recettes)
        self._apply_total_style(ws, row)
        row += 1
        ws.cell(row=row, column=4, value='Dépenses:')
        ws.cell(row=row, column=5, value=-total_depenses)
        self._apply_total_style(ws, row)
        row += 1
        ws.cell(row=row, column=4, value='Solde:')
        ws.cell(row=row, column=5, value=solde)
        self._apply_total_style(ws, row)
        
        self._auto_width(ws)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_invoices_excel(self, invoices, user_name=""):
        """Génère un fichier Excel de la liste des factures"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Factures"
        
        # Titre
        title = f"Factures de {user_name}" if user_name else "Liste des factures"
        self._add_title_row(ws, title, 5)
        
        # En-têtes
        headers = ['Numéro', 'Date', 'Client', 'Total (FCFA)', 'Statut']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        self._apply_header_style(ws, 4)
        
        # Données
        total_montant = 0
        row = 5
        
        for inv in invoices:
            total_montant += inv.total
            
            ws.cell(row=row, column=1, value=inv.numero)
            ws.cell(row=row, column=2, value=inv.date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=inv.client_name)
            ws.cell(row=row, column=4, value=inv.total)
            ws.cell(row=row, column=5, value=inv.status.capitalize())
            row += 1
        
        self._apply_data_style(ws, 5, row - 1)
        
        # Total
        ws.cell(row=row, column=3, value='TOTAL:')
        ws.cell(row=row, column=4, value=total_montant)
        self._apply_total_style(ws, row)
        
        self._auto_width(ws)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_invoice_detail_excel(self, invoice):
        """Génère un fichier Excel détaillé d'une facture"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Facture {invoice.numero}"
        
        # En-tête
        ws['A1'] = f"FACTURE {invoice.numero}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = 'Date:'
        ws['B3'] = invoice.date.strftime('%d/%m/%Y')
        ws['C3'] = 'Client:'
        ws['D3'] = invoice.client_name
        ws['A4'] = 'Statut:'
        ws['B4'] = invoice.status.capitalize()
        
        # Tableau articles
        headers = ['Produit', 'Quantité', 'Prix unitaire (FCFA)', 'Total (FCFA)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=6, column=col, value=header)
        self._apply_header_style(ws, 6)
        
        row = 7
        for item in invoice.items:
            ws.cell(row=row, column=1, value=item.product_name)
            ws.cell(row=row, column=2, value=item.quantity)
            ws.cell(row=row, column=3, value=item.unit_price)
            ws.cell(row=row, column=4, value=item.total)
            row += 1
        
        self._apply_data_style(ws, 7, row - 1)
        
        # Total
        ws.cell(row=row, column=3, value='TOTAL:')
        ws.cell(row=row, column=4, value=invoice.total)
        self._apply_total_style(ws, row)
        
        # Notes
        if invoice.notes:
            row += 2
            ws.cell(row=row, column=1, value='Notes:')
            ws.cell(row=row, column=2, value=invoice.notes)
        
        self._auto_width(ws)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
